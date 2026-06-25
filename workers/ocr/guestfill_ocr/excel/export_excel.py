"""Export OCR results to Excel."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from guestfill_ocr.excel.columns import GUEST_COLUMNS
from guestfill_ocr.excel.diagnostics_sheet import write_diagnostics_sheet
from guestfill_ocr.excel.errors_sheet import write_errors_sheet
from guestfill_ocr.excel.instructions_sheet import write_instructions
from guestfill_ocr.excel.styles import (
    style_confidence_cell,
    style_header,
    style_status_cell,
    style_warning_cell,
)
from guestfill_ocr.excel.validation_lists import add_filters, add_status_dropdown


def export_to_excel(
    rows: list[dict], errors: list[dict], diagnostics: list[dict], output_path: str, options: dict
) -> None:
    wb = Workbook()
    guests_ws = wb.active
    guests_ws.title = "Guests"

    _write_guest_headers(guests_ws)
    _write_guest_rows(guests_ws, rows)
    _style_guest_sheet(guests_ws, len(rows))

    if options.get("includeErrorsSheet", True):
        errors_ws = wb.create_sheet("Errors")
        write_errors_sheet(errors_ws, errors)

    if options.get("includeInstructionsSheet", True):
        instructions_ws = wb.create_sheet("Instructions")
        write_instructions(instructions_ws)

    if options.get("enableDiagnosticsSheet", True):
        diag_ws = wb.create_sheet("Diagnostics")
        write_diagnostics_sheet(diag_ws, diagnostics)

    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path_obj))


def _write_guest_headers(ws) -> None:
    for col_idx, col_name in enumerate(GUEST_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)


def _write_guest_rows(ws, rows: list[dict]) -> None:
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(GUEST_COLUMNS, 1):
            value = row_data.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)


def _style_guest_sheet(ws, num_rows: int) -> None:
    num_cols = len(GUEST_COLUMNS)
    style_header(ws, num_cols)

    status_col = GUEST_COLUMNS.index("status") + 1
    warning_col = GUEST_COLUMNS.index("ocr_warning") + 1
    confidence_col = GUEST_COLUMNS.index("confidence_score") + 1
    confidence_level_col = GUEST_COLUMNS.index("confidence_level") + 1

    for row_idx in range(2, num_rows + 2):
        status_cell = ws.cell(row=row_idx, column=status_col)
        style_status_cell(status_cell, str(status_cell.value or ""))

        warning_cell = ws.cell(row=row_idx, column=warning_col)
        style_warning_cell(warning_cell, str(warning_cell.value or ""))

        conf_level_cell = ws.cell(row=row_idx, column=confidence_level_col)
        style_confidence_cell(conf_level_cell, str(conf_level_cell.value or ""))

    add_status_dropdown(ws, num_rows, get_column_letter(status_col))
    add_filters(ws, num_cols, num_rows)

    ws.freeze_panes = "A2"
