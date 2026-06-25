"""Excel styling utilities."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, num_cols: int, max_width: int = 40) -> None:
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max_len + 3, max_width)
        ws.column_dimensions[get_column_letter(col)].width = max(adjusted, 10)


def style_status_cell(cell, status: str) -> None:
    if status == "READY":
        cell.fill = GREEN_FILL
    elif status == "NEED_REVIEW":
        cell.fill = YELLOW_FILL
    elif status == "FAILED":
        cell.fill = RED_FILL


def style_warning_cell(cell, value: str) -> None:
    if value:
        cell.fill = YELLOW_FILL


def style_confidence_cell(cell, level: str) -> None:
    if level == "LOW":
        cell.fill = RED_FILL
    elif level == "MEDIUM":
        cell.fill = YELLOW_FILL
    elif level == "HIGH":
        cell.fill = GREEN_FILL
